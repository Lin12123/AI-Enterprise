"""知识导入解析。

分级策略(对齐 docs/cloud_platform_design.md §5.4)：
- Excel(.xlsx)/JSON  → 结构化明确，高置信，直接产出 published 条目
- Word/PDF/图片       → 需抽取+人工确认，先产出 draft(此处留 TODO)

返回统一的中间结构 list[ParsedRule]，由路由层落库。
"""
import json
from typing import Any


def parse_json(content: bytes) -> list[dict[str, Any]]:
    """JSON 导入：期望 [{scope_material, scope_feature, clause, params_json}, ...]。"""
    data = json.loads(content.decode("utf-8"))
    if isinstance(data, dict):
        data = data.get("rules", [])
    rules: list[dict[str, Any]] = []
    for item in data:
        rules.append({
            "scope_material": item.get("scope_material"),
            "scope_feature": item.get("scope_feature"),
            "clause": item.get("clause"),
            "params_json": item.get("params_json") or {},
            "status": "published",
        })
    return rules


def parse_xlsx(abs_path: str) -> list[dict[str, Any]]:
    """Excel 导入：首行表头，列名 scope_material/scope_feature/clause/params_json(JSON字符串)。"""
    from openpyxl import load_workbook  # 延迟导入，未装依赖时不影响其它端点

    wb = load_workbook(abs_path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    header = [str(h).strip() if h is not None else "" for h in rows[0]]
    idx = {name: i for i, name in enumerate(header)}
    out: list[dict[str, Any]] = []
    for row in rows[1:]:
        if row is None or all(c is None for c in row):
            continue

        def cell(name: str):
            i = idx.get(name)
            return row[i] if i is not None and i < len(row) else None

        params_raw = cell("params_json")
        params: dict[str, Any] = {}
        if params_raw:
            try:
                params = json.loads(str(params_raw))
            except (ValueError, TypeError):
                params = {"raw": str(params_raw)}
        out.append({
            "scope_material": cell("scope_material"),
            "scope_feature": cell("scope_feature"),
            "clause": cell("clause"),
            "params_json": params,
            "status": "published",
        })
    return out


def parse_document(fmt: str, abs_path: str) -> list[dict[str, Any]]:
    """Word/PDF：抽取正文文本 → 本地 LLM 理解 → 产出 status='draft' 规则草稿。

    图片(png/jpg)本轮暂缓 OCR，返回空列表(仅存原文附件、待人工补录)。
    LLM 不可达或抽取失败时同样返回空列表，交由路由层降级提示。
    """
    from . import doc_extractor, rule_extractor  # 延迟导入，避免未装依赖影响其它端点

    if not doc_extractor.is_text_extractable(fmt):
        # 图片等：暂不支持自动抽取
        return []
    try:
        text = doc_extractor.extract_text(fmt, abs_path)
    except doc_extractor.DocExtractError:
        return []
    if not text:
        return []
    return rule_extractor.extract_rules_from_text(text)


def dispatch(fmt: str, abs_path: str, content: bytes) -> tuple[list[dict[str, Any]], bool]:
    """按格式分派，返回 (解析条目, 是否高置信可直接入库)。"""
    fmt = (fmt or "").lower().lstrip(".")
    if fmt == "json":
        return parse_json(content), True
    if fmt in ("xlsx", "xls"):
        return parse_xlsx(abs_path), True
    if fmt in ("docx", "pdf", "png", "jpg", "jpeg"):
        # 文档/图片：LLM 抽取草稿，低置信，需人工确认后发布
        return parse_document(fmt, abs_path), False
    return [], False